from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pytest
from openpyxl import Workbook, load_workbook


@pytest.fixture()
def log():
    print("testFixture")
    return "value"

def test_first(log):

    book = load_workbook('ConnAutoTest.xlsx')

    sheet = book.active

    print(sheet['I3'].value)

    row = sheet.max_row

    print(f'{row} строк')



    driver = webdriver.Chrome()

    driver.maximize_window()

    driver.get("https://staging.connectable.site/login")

    try:

        driver.find_element(By.XPATH, '//input[@placeholder="Workspace"]').send_keys('testing3')

        driver.find_element(By.XPATH, '//input[@placeholder="E-mail"]').send_keys('t2@gmail.com')

        driver.find_element(By.XPATH, '//input[@placeholder="Password"]').send_keys('111111')

        driver.find_element(By.XPATH, '//div[text()="Log in"]').click()

        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//div[text()=' Администрирование ']")))

        driver.find_element(By.XPATH, "//div[text()=' Администрирование ']").click()

        for i in range(2, row):

            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//div[text()='Новый сотрудник']")))

            driver.find_element(By.XPATH, "//div[text()='Новый сотрудник']").click()

            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//div[@class='ant-modal-body']")))

            driver.find_element(By.XPATH, "//input[@placeholder='Введите имя сотрудника']").send_keys(
                sheet[f'B{i}'].value)
            driver.find_element(By.XPATH, "//input[@placeholder='Введите отчество сотрудника']").send_keys(
                sheet[f'C{i}'].value)
            driver.find_element(By.XPATH, "//input[@placeholder='Введите фамилию сотрудника']").send_keys(
                sheet[f'A{i}'].value)
            time.sleep(4)
            #Выводится не правильный формат даты из экселя
            driver.find_element(By.XPATH, "//input[@placeholder='Выберите дату']").send_keys(
                sheet[f'I{i}'].value)
            driver.find_element(By.XPATH, "//input[@placeholder='Должность']").send_keys(
                sheet[f'E{i}'].value)
            driver.find_element(By.XPATH, "(//input[@placeholder='Телефон'])[1]").send_keys(
                sheet[f'G{i}'].value)
            driver.find_element(By.XPATH, "(//input[@placeholder='Телефон'])[2]").send_keys(
                sheet[f'F{i}'].value)
            driver.find_element(By.XPATH, "//input[@placeholder='Кабинет']").send_keys(
                sheet[f'H{i}'].value)
            driver.find_element(By.XPATH, "//input[@placeholder='Введите e-mail']").send_keys(
                sheet[f'D{i}'].value)
            driver.find_element(By.XPATH, "//input[@placeholder='Пароль']").send_keys('111111')
            driver.find_element(By.XPATH, "//input[@placeholder='Подтверждение пароля']").send_keys('111111')
            time.sleep(1)
            driver.find_element(By.XPATH, "(//button[@class='ant-btn ant-btn-primary'])[2]").click()
            time.sleep(3)


    except:
        print('ERROR')
    finally:
        driver.quit()