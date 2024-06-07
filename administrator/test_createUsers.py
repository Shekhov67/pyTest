from selenium import webdriver
from selenium.common import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pytest
from openpyxl import Workbook, load_workbook

@pytest.fixture()
def page():
    ''' Переход на страницу портала '''
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://staging.connectable.site/")
    #driver.get("https://connectable.site/")
    driver.implicitly_wait(5)
    return driver



def test_first(page):
    '''page это драйвер с уже запущенной старницей'''

    book = load_workbook('../Data files/ConnAutoTestClient.xlsx')

    sheet = book.active

    row = sheet.max_row

    print(f'{row} строк')

    page.find_element(By.XPATH, '//input[@placeholder="Workspace"]').send_keys('pytest')

    page.find_element(By.XPATH, '//input[@placeholder="E-mail"]').send_keys('t2@gmail.com')

    page.find_element(By.XPATH, '//input[@placeholder="Password"]').send_keys('111111')

    page.find_element(By.XPATH, '//div[text()="Log in"]').click()

    WebDriverWait(page, 5).until(
        EC.presence_of_element_located((By.XPATH, "//*[contains(text(), ' Администрирование')]")))

    page.find_element(By.XPATH, "//*[contains(text(), ' Администрирование')]").click()

        # row - количество заполненых строк в эксель файле, от количества строк зависит количество циклов
    for i in range(2, row+1):

        WebDriverWait(page, 5).until(
            EC.presence_of_element_located((By.XPATH, "//div[text()='Новый сотрудник']")))

        page.find_element(By.XPATH, "//div[text()='Новый сотрудник']").click()

        WebDriverWait(page, 5).until(
            EC.presence_of_element_located((By.XPATH, "//div[@class='ant-modal-body']")))

        if sheet[f'B{i}'].value is None:
            print("not name")
        else:
            page.find_element(By.XPATH, "//input[@placeholder='Введите имя сотрудника']").send_keys(
                sheet[f'B{i}'].value)

        if sheet[f'C{i}'].value is None:
            print("not patronymic")
        else:
            page.find_element(By.XPATH, "//input[@placeholder='Введите отчество сотрудника']").send_keys(
                sheet[f'C{i}'].value)
        if sheet[f'A{i}'].value is None:
            print("not surname")
        else:
            page.find_element(By.XPATH, "//input[@placeholder='Введите фамилию сотрудника']").send_keys(
                sheet[f'A{i}'].value)
            #Формат даты из экселя перевожу в строку
            # excel_date берет дату из файла эксель в формате 1994-01-21 00:00:00
        if sheet[f'I{i}'].value is None:
            print("not date birth")
        else:
            # new_date - переводит формат даты с 1994-01-21 00:00:00 в 22.01.1994, сделано для того,
            # чтобы добавить дату рождения при создании сотрудника
            excel_date = sheet[f'I{i}'].value
            new_date = excel_date.strftime('%d.%m.%Y')
            page.find_element(By.XPATH, "//input[@placeholder='Выберите дату']").send_keys(new_date)

        if sheet[f'E{i}'].value is None:
            print("not position")
        else:
            page.find_element(By.XPATH, "//input[@placeholder='Должность']").send_keys(
                sheet[f'E{i}'].value)

        if sheet[f'G{i}'].value is None:
            print("not phone")
        else:
            page.find_element(By.XPATH, "(//input[@placeholder='Телефон'])[1]").send_keys(
                sheet[f'G{i}'].value)
        if sheet[f'F{i}'].value is None:
            print("not phone2")
        else:
            page.find_element(By.XPATH, "(//input[@placeholder='Телефон'])[2]").send_keys(
                sheet[f'F{i}'].value)
        if sheet[f'H{i}'].value is None:
            print("not cabinet")
        else:
            page.find_element(By.XPATH, "//input[@placeholder='Кабинет']").send_keys(
                sheet[f'H{i}'].value)
        if sheet[f'D{i}'].value is None:
            print("not email")
        else:
            page.find_element(By.XPATH, '//input[@placeholder="Введите e-mail"]').send_keys(
                sheet[f'D{i}'].value)
        page.find_element(By.XPATH, "//input[@placeholder='Пароль']").send_keys('111111')
        page.find_element(By.XPATH, "//input[@placeholder='Подтверждение пароля']").send_keys('111111')
        page.find_element(By.XPATH, "(//button[@class='ant-btn ant-btn-primary'])[2]").click()

        WebDriverWait(page, 5).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@class= 'anticon anticon-close ant-notification-close-icon']")))
        page.find_element(By.XPATH, "//*[@class= 'anticon anticon-close ant-notification-close-icon']").click()


""""
        try:
            WebDriverWait(page, 5).until(EC.element_to_be_clickable((By.XPATH, "//span[text()= 'Ошибка при сохранении сотрудника']")))
            page.find_element(By.XPATH, "//span[text()= 'Ошибка при сохранении сотрудника']")
            errText = page.find_element(By.XPATH, '//div[@class="ant-modal-confirm-content"]').text
            print(errText)
            WebDriverWait(page, 5).until(EC.element_to_be_clickable((By.XPATH, '(//button[@class="ant-btn ant-btn-primary"])[3]')))
            page.find_element(By.XPATH, '(//button[@class="ant-btn ant-btn-primary"])[3]').click()
            WebDriverWait(page, 5).until(EC.element_to_be_clickable((By.XPATH, '//button[@class="ant-btn"]')))
            page.find_element(By.XPATH, '//button[@class="ant-btn"]').click()
        except (TimeoutException, NoSuchElementException):
            page.find_element(By.XPATH, "//span[text()= 'Ошибка при сохранении сотрудника']").is_enabled()
            print(f"Сотрудник {i} создан")
            WebDriverWait(page, 5).until(EC.element_to_be_clickable(
                (By.XPATH, "//*[@class= 'anticon anticon-close ant-notification-close-icon']")))
            page.find_element(By.XPATH, "//*[@class= 'anticon anticon-close ant-notification-close-icon']").click()

"""